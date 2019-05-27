# -*- coding: utf-8 -*-

import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import csv, operator, openpyxl
from xlrd import open_workbook

class Alumno:
    def __init__(self, num_cuenta, nombre, appaterno, apmaterno, cal1, cal2, cal3, cal4, cal5):
        self.num_cuenta = num_cuenta
        self.nombre = nombre
        self.appaterno = appaterno
        self.apmaterno = apmaterno
        self.cal1 = cal1
        self.cal2 = cal2
        self.cal3 = cal3
        self.cal4 = cal4
        self.cal5 = cal5
        self.promedio = 0
    
    # Calcular promedio
    def calcular_promedio(self):
        listLength = len (datos) -4
        self.promedio = (self.cal1 + self.cal2 + self.cal3 + self.cal4 + self.cal5)/listLength
        return self.promedio
    
    def cadena(self):
        datos = [self.num_cuenta, self.nombre, self.appaterno, self.apmaterno, self.cal1, self.cal2, self.cal3, self.cal4, self.cal5, alumno.calcular_promedio()]
        return datos
    
    
    def __str__(self): #Regresa cadena, todos los objetos tiene su metodo str
        return "Num_cuenta: " + self.num_cuenta + "\nNombre: " + self.nombre + " " + self.appaterno + " " + self.apmaterno + "\nPromedio: " + str(alumno.calcular_promedio())

def read_promedios():
    try:
        wb= openpyxl.load_workbook('Promedios.xlsx')
        sheet = wb["Sheet1"]
        #Identificamos los máximos y mínimos de la hoja
        min_row = sheet.min_row + 1
        max_row = sheet.max_row
        max_col = sheet.max_column
        min_col = sheet.min_column

        for x in range(min_row, max_row+1, 1):
            print ("--------------------------")
            for y in range(min_col, max_col+1, 1):
                celdas = ["","A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
                var = celdas[y] + str(x)
                if celdas[y] == "A":
                    print (var + " \tNum_Cuenta:\t" + str(sheet[var].value))
                elif celdas[y] == "B":
                    print (var + " \tNombre:\t" + str(sheet[var].value))
                elif celdas[y] == "C":
                    print (var + " \tAp_Paterno:\t" + str(sheet[var].value))
                elif celdas[y] == "D":
                    print (var + " \tAp_Materno:\t" + str(sheet[var].value))
                elif celdas[y] == "E":
                    print (var + " \tCalificación 1:\t" + str(sheet[var].value))
                elif celdas[y] == "F":
                    print (var + " \tCalificación 2:\t" + str(sheet[var].value))
                elif celdas[y] == "G":
                    print (var + " \tCalificación 3:\t" + str(sheet[var].value))
                elif celdas[y] == "H":
                    print (var + " \tCalificación 4:\t" + str(sheet[var].value))
                elif celdas[y] == "I":
                    print (var + " \tCalificación 5:\t" + str(sheet[var].value))
                elif celdas[y] == "J":
                    print (var + " \tPromedio:\t" + str(sheet[var].value))
    except:
        print("Error en lectura")
        
def write_promedios(data):
    try:
        wb= openpyxl.load_workbook('Promedios.xlsx')
        sheet = wb["Sheet1"]
        print (data)
        #Identificamos los máximos y mínimos de la hoja
        min_row = sheet.min_row + 1
        max_row = sheet.max_row
        max_col = sheet.max_column
        min_col = sheet.min_column
        print (min_row)
        print (max_row)
        print (max_col)
        print (min_col)
        print (sheet.rows)
        for x in range(max_row, max_row+1, 1):
            print ("--------------------------")
            for y in range(min_col, max_col +1, 1):
                celdas = ["","A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
                for z in data:
                    var = celdas[y] + str(x+1)
                    sheet[var] = data[y-1]
                    wb.save("Promedios.xlsx")
    except:
        print("Error en escritura")

def obtener_data():
    try:
        wb= openpyxl.load_workbook('Promedios.xlsx')
        sheet = wb["Sheet1"]
        #Identificamos los máximos y mínimos de la hoja
        min_row = sheet.min_row + 1
        max_row = sheet.max_row
        max_col = sheet.max_column
        min_col = sheet.min_column
        
        lista_generica = list(range(max_row-1))
        lista_a = list(range(max_row-1)) 
        lista_b = list(range(max_row-1))
        lista_e = list(range(max_row-1))
        lista_f = list(range(max_row-1))
        lista_g = list(range(max_row-1))
        lista_h = list(range(max_row-1))
        lista_i = list(range(max_row-1))
        lista_j = list(range(max_row-1))
        lista_w = list(range(max_row-1))
        z=0
        z0=0
        z1=0
        z2=0
        z3=0
        z4=0
        z5=0
        z6=0
        z7=0
        
        for x in range(min_row, max_row+1, 1):
            for y in range(min_col, max_col+1, 1):
            #for y in range(min_col, max_col+1, 1):
                celdas = ["","A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
                var = celdas[y] + str(x)
                if celdas[y] == "A":
                    lista_a[z] = sheet[var].value
                    z = z +1
                if celdas[y] == "B":
                    lista_b[z0] = sheet[var].value
                    z0 = z0 +1
                if celdas[y] == "E":
                    lista_e[z1] = sheet[var].value
                    z1 = z1 +1
                if celdas[y] == "F":
                    lista_f[z2] = sheet[var].value
                    z2 = z2 +1
                if celdas[y] == "G":
                    lista_g[z3] = sheet[var].value
                    z3 = z3 +1
                if celdas[y] == "H":
                    lista_h[z4] = sheet[var].value
                    z4 = z4 +1
                if celdas[y] == "I":
                    lista_i[z5] = sheet[var].value
                    z5 = z5 +1
                if celdas[y] == "J":
                    lista_j[z6] = sheet[var].value
                    z6 = z6 +1
        #print ("Info")
        info = [lista_a, lista_b]
        #print (info)
        #print ("Data")
        data = [lista_e, lista_f, lista_g, lista_h, lista_i, lista_j]
        #print (data)
        axisX = np.arange(len(data[0])) 
        print(axisX+.10)
        plt.bar(axisX + 0.0, data[0], color = "b", width = 0.10)
        plt.bar(axisX + 0.1, data[1], color = "c", width = 0.10)
        plt.bar(axisX + 0.2, data[2], color = "r", width = 0.10)
        plt.bar(axisX + 0.3, data[3], color = "g", width = 0.10)
        plt.bar(axisX + 0.4, data[4], color = "m", width = 0.10)
        plt.bar(axisX + 0.5, data[5], color = "y", width = 0.10)
        plt.xticks(axisX+0.1, lista_b)
        nom = "graph.png"
        plt.savefig(nom)
        plt.show()
    except:
        print("Error en escritura")

running = True
while running:
    valor_1 = 0
    valor_2 = 0
    print ("---Elige una opción---")
    print ("1- Agregar registro de alumno")
    print ("2- Leer Datos")
    print ("3- Graficar")
    print ("4- Salir")
    op = int(input('Opcion: '))
    if op == 1:
        print ("---Agrega registros---")
        print ("Ingresa número de cuenta:")
        num_cuenta = input(' ')    
        print ("Ingresa nombre:")
        nom = input(' ')  
        print ("Ingresa apellido paterno:")
        app = input(' ')
        print ("Ingresa apellido materno:")
        apm = input(' ')  
        print ("Ingresa primera calificación:")
        c1 = int(input(' '))
        print ("Ingresa segunda calificación:")
        c2 = int(input(' '))
        print ("Ingresa tercera calificación:")
        c3 = int(input(' '))
        print ("Ingresa cuarta calificación:")
        c4 = int(input(' '))
        print ("Ingresa quinta calificación:")
        c5 = int(input(' '))
        datos = [num_cuenta, nom, app, apm, c1, c2, c3, c4, c5]
        print (datos);
        print ("Alumno:")
        alumno = Alumno(datos[0], datos[1], datos[2], datos[3], datos[4], datos[5], datos[6], datos[7], datos[8])
        #print(alumno)
        print("Datos:")
        print (alumno.cadena())
        data = alumno.cadena()
        write_promedios(data)
    elif op == 2:
        read_promedios()
    elif op == 3:
        obtener_data()
    elif op >= 4:
        print ("Bye")
        running = False
